import openpyxl as xl
path = "C:/Users/User/Desktop/xxx/TestTask.xlsx"
workbook = xl.load_workbook(path)
sheet = workbook.active

rows = sheet.max_row
column = sheet.max_column
#print(rows,column)
for r in range(1,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value,end="       ")
    print()